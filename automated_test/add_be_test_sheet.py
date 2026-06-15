#!/usr/bin/env python3
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from export_dast_xlsx import load_rows


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = ROOT / "E2E_Test_Report_MyJoints_Latest.xlsx"


def styles():
    return {
        "header_fill": PatternFill(start_color="008080", end_color="008080", fill_type="solid"),
        "summary_bg": PatternFill(start_color="E6F2F2", end_color="E6F2F2", fill_type="solid"),
        "passed_fill": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        "failed_fill": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        "font_bold_white": Font(name="Arial", size=11, bold=True, color="FFFFFF"),
        "font_bold": Font(name="Arial", size=11, bold=True),
        "font_regular": Font(name="Arial", size=10),
        "font_title": Font(name="Arial", size=16, bold=True, color="008080"),
        "thin_border": Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        ),
    }


def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    if "BE TEST" in wb.sheetnames:
        del wb["BE TEST"]

    ws = wb.create_sheet("BE TEST")
    ws.views.sheetView[0].showGridLines = True
    st = styles()
    rows = load_rows()
    findings = [row for row in rows if row["finding"]]
    counts = Counter(row["severity"] for row in findings)

    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 42
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 55

    ws.append([])
    ws.append(["", "My Joints Backend DAST Results"])
    ws["B2"].font = st["font_title"]

    ws.append([])
    ws.append(["", "Metric", "Value"])
    for coord in ("B4", "C4"):
        ws[coord].font = st["font_bold_white"]
        ws[coord].fill = st["header_fill"]

    metrics = [
        ("Total Backend Checks", len(rows)),
        ("Backend Findings", len(findings)),
        ("Critical Findings", counts.get("critical", 0)),
        ("High Findings", counts.get("high", 0)),
        ("Medium Findings", counts.get("medium", 0)),
        ("Included Source", "automated_test/report.json"),
    ]
    for metric, value in metrics:
        ws.append(["", metric, value])

    for row in range(5, 5 + len(metrics)):
        ws.cell(row=row, column=2).font = st["font_bold"]
        ws.cell(row=row, column=2).fill = st["summary_bg"]
        ws.cell(row=row, column=2).border = st["thin_border"]
        ws.cell(row=row, column=3).font = st["font_regular"]
        ws.cell(row=row, column=3).border = st["thin_border"]

    start = 12
    headers = [
        "Test Case ID",
        "Test Type",
        "Severity",
        "Endpoint",
        "Expected Result",
        "Status",
        "Duration (s)",
        "Notes / Errors",
    ]
    for idx, header in enumerate(headers, start=2):
        cell = ws.cell(row=start, column=idx)
        cell.value = header
        cell.font = st["font_bold_white"]
        cell.fill = st["header_fill"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = st["thin_border"]

    for row_idx, row in enumerate(rows, start=start + 1):
        status_text = "Failed" if row["finding"] else "Passed"
        values = [
            f"BE-{row_idx - start:03d}",
            row["test_category"],
            row["severity"].upper(),
            f"{row['method']} {row['endpoint']}",
            f"Expected {row['expected_status']}; got {row['status']}",
            status_text,
            round((row.get("response_time_ms") or 0) / 1000, 3),
            f"role={row['role']} | {row['note']}",
        ]
        for col_idx, value in enumerate(values, start=2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = st["font_regular"]
            cell.border = st["thin_border"]
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        status_cell = ws.cell(row=row_idx, column=7)
        if row["finding"]:
            status_cell.fill = st["failed_fill"]
            status_cell.font = Font(name="Arial", size=10, color="721C24", bold=True)
        else:
            status_cell.fill = st["passed_fill"]
            status_cell.font = Font(name="Arial", size=10, color="155724", bold=True)

    wb.save(WORKBOOK_PATH)
    print(f"Updated {WORKBOOK_PATH} with BE TEST sheet")


if __name__ == "__main__":
    main()
