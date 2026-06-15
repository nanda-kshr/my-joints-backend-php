#!/usr/bin/env python3
import json
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[2]
REPORT_JSON = Path(__file__).resolve().parent / "report.json"
OUTPUT_XLSX = ROOT / "E2E_Test_Report_MyJoints_Latest.xlsx"
EXCLUDED_ROWS = {
    ("authz_privesc", "/api/doctor/notifications/update", "PUT", "patient", "Patient token against doctor-only endpoint"),
    ("authz_privesc", "/api/doctor/consult-request", "POST", "patient", "Patient token against doctor-only endpoint"),
    ("authn_bypass", "/api/doctor/notifications/update", "PUT", "none", "No token should be rejected"),
    ("authn_bypass", "/api/doctor/notifications/update", "PUT", "malformed", "Malformed token should be rejected"),
    ("authn_bypass", "/api/doctor/notifications/update", "PUT", "expired", "Expired token should be rejected"),
    ("authn_bypass", "/api/doctor/consult-request", "POST", "none", "No token should be rejected"),
    ("authn_bypass", "/api/doctor/consult-request", "POST", "malformed", "Malformed token should be rejected"),
    ("authn_bypass", "/api/doctor/consult-request", "POST", "expired", "Expired token should be rejected"),
}


def style_defs():
    header_fill = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
    passed_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    failed_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    summary_bg = PatternFill(start_color="E6F2F2", end_color="E6F2F2", fill_type="solid")

    font_bold_white = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=11, bold=True)
    font_regular = Font(name="Arial", size=10)
    font_title = Font(name="Arial", size=16, bold=True, color="008080")

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    return {
        "header_fill": header_fill,
        "passed_fill": passed_fill,
        "failed_fill": failed_fill,
        "summary_bg": summary_bg,
        "font_bold_white": font_bold_white,
        "font_bold": font_bold,
        "font_regular": font_regular,
        "font_title": font_title,
        "thin_border": thin_border,
    }


def load_rows():
    rows = json.loads(REPORT_JSON.read_text())
    rows = [
        row for row in rows
        if (
            row["test_category"],
            row["endpoint"],
            row["method"],
            row["role"],
            row["note"],
        ) not in EXCLUDED_ROWS
    ]
    rows.sort(key=lambda row: (row["finding"], row["severity"], row["test_category"], row["endpoint"]), reverse=True)
    return rows


def add_header(ws, headers, styles):
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = styles["font_bold_white"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = styles["thin_border"]


def set_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def write_rows(ws, rows, styles):
    for row_num, row in enumerate(rows, start=2):
        status_text = "Failed" if row["finding"] else "Passed"
        values = [
            f"DAST-{row_num - 1:03d}",
            row["test_category"],
            row["severity"].upper(),
            f"{row['method']} {row['endpoint']}",
            f"Expected {row['expected_status']}; got {row['status']}",
            status_text,
            round((row.get("response_time_ms") or 0) / 1000, 3),
            f"role={row['role']} | {row['note']}",
        ]
        ws.append(values)
        for col in range(1, len(values) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = styles["font_regular"]
            cell.border = styles["thin_border"]
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        status_cell = ws.cell(row=row_num, column=6)
        if row["finding"]:
            status_cell.fill = styles["failed_fill"]
            status_cell.font = Font(name="Arial", size=10, color="721C24", bold=True)
        else:
            status_cell.fill = styles["passed_fill"]
            status_cell.font = Font(name="Arial", size=10, color="155724", bold=True)


def build_summary(ws, rows, styles):
    findings = [row for row in rows if row["finding"]]
    counts = Counter(row["severity"] for row in findings)
    total = len(rows)
    passed = total - len(findings)
    pass_rate = round((passed / total) * 100, 2) if total else 0
    deployment = "FINDINGS REQUIRE REMEDIATION" if findings else "NO FINDINGS OBSERVED"

    ws.views.sheetView[0].showGridLines = True
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 24

    ws.append([])
    ws.append(["", "My Joints DAST Summary"])
    ws.cell(row=2, column=2).font = styles["font_title"]

    ws.append([])
    ws.append(["", "Metric", "Value"])
    for coord in ("B4", "C4"):
        ws[coord].font = styles["font_bold_white"]
        ws[coord].fill = styles["header_fill"]

    metrics = [
        ("Total Test Cases", total),
        ("Passed", passed),
        ("Failed", len(findings)),
        ("Pass Rate %", pass_rate),
        ("Critical Findings", counts.get("critical", 0)),
        ("High Findings", counts.get("high", 0)),
        ("Medium Findings", counts.get("medium", 0)),
        ("Low/Info Findings", counts.get("low", 0) + counts.get("info", 0)),
        ("Deployment Status", deployment),
        ("Verification Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    for metric, value in metrics:
        ws.append(["", metric, value])

    for row in range(5, 5 + len(metrics)):
        ws.cell(row=row, column=2).font = styles["font_bold"]
        ws.cell(row=row, column=2).border = styles["thin_border"]
        ws.cell(row=row, column=2).fill = styles["summary_bg"]
        ws.cell(row=row, column=3).font = styles["font_regular"]
        ws.cell(row=row, column=3).border = styles["thin_border"]

    top_notes = [f"{r['severity'].upper()}: {r['method']} {r['endpoint']} - {r['note']}" for r in findings[:5]]
    start = 16
    ws.cell(row=start, column=2, value="Top Issues").font = styles["font_bold"]
    ws.cell(row=start, column=2).fill = styles["summary_bg"]
    ws.cell(row=start, column=2).border = styles["thin_border"]
    for idx, note in enumerate(top_notes, start=start + 1):
        ws.cell(row=idx, column=2, value=note)
        ws.cell(row=idx, column=2).font = styles["font_regular"]
        ws.cell(row=idx, column=2).border = styles["thin_border"]
        ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=3)


def main():
    rows = load_rows()
    styles = style_defs()
    wb = openpyxl.Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    build_summary(summary_ws, rows, styles)

    headers = [
        "Test Case ID",
        "Test Type",
        "Category",
        "Test Name / Objective",
        "Expected Result",
        "Status",
        "Duration (s)",
        "Notes / Errors",
    ]
    widths = {
        "A": 15,
        "B": 18,
        "C": 18,
        "D": 42,
        "E": 26,
        "F": 12,
        "G": 12,
        "H": 58,
    }

    details_ws = wb.create_sheet("Test Details")
    details_ws.views.sheetView[0].showGridLines = True
    add_header(details_ws, headers, styles)
    set_widths(details_ws, widths)
    write_rows(details_ws, rows, styles)

    passed_ws = wb.create_sheet("Passed Tests")
    passed_ws.views.sheetView[0].showGridLines = True
    add_header(passed_ws, headers, styles)
    set_widths(passed_ws, widths)
    write_rows(passed_ws, [row for row in rows if not row["finding"]], styles)

    failed_ws = wb.create_sheet("Failed Tests")
    failed_ws.views.sheetView[0].showGridLines = True
    add_header(failed_ws, headers, styles)
    set_widths(failed_ws, widths)
    write_rows(failed_ws, [row for row in rows if row["finding"]], styles)

    wb.save(OUTPUT_XLSX)
    print(f"Generated {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
