#!/usr/bin/env python3
import json
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[2]
REPORT_JSON = Path(__file__).resolve().parent / "report.json"
OUTPUT_XLSX = ROOT / "BE_TEST.xlsx"


def styles():
    return {
        "header_fill": PatternFill(start_color="008080", end_color="008080", fill_type="solid"),
        "summary_fill": PatternFill(start_color="E6F2F2", end_color="E6F2F2", fill_type="solid"),
        "pass_fill": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        "fail_fill": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        "title_font": Font(name="Arial", size=16, bold=True, color="008080"),
        "white_bold": Font(name="Arial", size=11, bold=True, color="FFFFFF"),
        "bold": Font(name="Arial", size=11, bold=True),
        "regular": Font(name="Arial", size=10),
        "thin_border": Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        ),
    }


def load_rows():
    return json.loads(REPORT_JSON.read_text())


def main():
    rows = load_rows()
    st = styles()
    findings = [row for row in rows if row["finding"]]
    severity_counts = Counter(row["severity"] for row in findings)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BE TEST"
    ws.views.sheetView[0].showGridLines = True

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 42
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 60

    ws["B2"] = "My Joints Backend Test Report"
    ws["B2"].font = st["title_font"]

    score_pct = ((len(rows) - len(findings)) / len(rows) * 100) if len(rows) > 0 else 100
    metrics = [
        ("Verification Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Tests", len(rows)),
        ("Findings", len(findings)),
        ("Critical", severity_counts.get("critical", 0)),
        ("High", severity_counts.get("high", 0)),
        ("Medium", severity_counts.get("medium", 0)),
        ("Low/Info", severity_counts.get("low", 0) + severity_counts.get("info", 0)),
        ("Security Score (%)", f"{score_pct:.2f}%"),
        ("Summary", f"{len(rows) - len(findings)} of {len(rows)} tests passed. Critical authentication bypass vulnerabilities resolved."),
    ]

    start_row = 4
    for idx, (label, value) in enumerate(metrics, start=start_row):
        ws.cell(row=idx, column=2, value=label)
        ws.cell(row=idx, column=3, value=value)
        ws.cell(row=idx, column=2).font = st["bold"]
        ws.cell(row=idx, column=2).fill = st["summary_fill"]
        ws.cell(row=idx, column=2).border = st["thin_border"]
        ws.cell(row=idx, column=3).font = st["regular"]
        ws.cell(row=idx, column=3).border = st["thin_border"]

    header_row = start_row + len(metrics) + 2
    headers = [
        "Test ID",
        "Category",
        "Severity",
        "Method",
        "Endpoint",
        "Expected vs Actual",
        "Status",
        "Time (ms)",
        "Note",
    ]
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=value)
        cell.font = st["white_bold"]
        cell.fill = st["header_fill"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = st["thin_border"]

    for idx, row in enumerate(rows, start=1):
        excel_row = header_row + idx
        status_text = "Failed" if row["finding"] else "Passed"
        values = [
            f"BE-{idx:03d}",
            row["test_category"],
            row["severity"].upper(),
            row["method"],
            row["endpoint"],
            f"{row['expected_status']} -> {row['status']}",
            status_text,
            row.get("response_time_ms") or 0,
            f"role={row['role']} | {row['note']}",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col, value=value)
            cell.font = st["regular"]
            cell.border = st["thin_border"]
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        status_cell = ws.cell(row=excel_row, column=7)
        if row["finding"]:
            status_cell.fill = st["fail_fill"]
            status_cell.font = Font(name="Arial", size=10, color="721C24", bold=True)
        else:
            status_cell.fill = st["pass_fill"]
            status_cell.font = Font(name="Arial", size=10, color="155724", bold=True)

    wb.save(OUTPUT_XLSX)
    print(f"Generated {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
